#upload xlsx name : email to DB
import pymysql
# 엑셀 다루기
from openpyxl import load_workbook

# 엑셀파일에서 이름 메일 읽어오기
# wb = load_workbook('target.xlsx', data_only=True, read_only=True)
wb = load_workbook('target.xlsx', data_only=True, read_only=True)
ws = wb.active

target = {}

#엑셀파일 파싱 reciever 디렉터리에 이름 : 메일 형태로 집어 넣음
for col in ws.iter_rows(min_row=2):
    index = 0
    for cell in col:
        if (cell.value == None):
            pass
        elif(index == 0):
            # 첫번째 나오는아이 -> 이름 (엑셀에서 이름, 메일 순서로 받아옴) -> 딕셔너리 key로 만듬
            tmp_key = cell.value
        else:
            # 두번째 나오는 아이 -> 메일 -> 딕셔너리에 value로 들어감
            target[tmp_key] = cell.value
        index = index+1

#DB연결
connection = pymysql.connect(
    host='localhost',
    # host='155.230.52.54',
    user='ksj',
    password='qhdks!321',
    db='WebProject',
    charset='utf8'
)

try:
    cursor = connection.cursor(pymysql.cursors.DictCursor)
    sql = "insert ignore into import_company(no, name, email)" \
        "values (%s, %s, %s)"
    no = 1
    for name, mail in target.items():
        cursor.execute(sql, (str(no), name, mail))
        no=no+1

    connection.commit()

finally:
    connection.close()